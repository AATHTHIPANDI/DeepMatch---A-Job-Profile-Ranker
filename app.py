import os
import json
import streamlit as st
import pandas as pd

# Import our custom modules
from src.extractor import extract_text_from_file
from src.jd_parser import parse_job_description
from src.candidate_profile import parse_raw_resume_text, synthesize_candidate_profile
from src.scorer import score_candidate
from src.explainability import generate_candidate_rationale

def generate_excel_report(ranked_list):
    import io
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Create list of dicts
    data = []
    for rank, cand in enumerate(ranked_list):
        sub = cand.get("sub_scores", {})
        rat = cand.get("rationale", {})
        snap = cand.get("candidate_profile_snapshot", {})
        
        # Format strengths and gaps
        strengths = "\n".join([f"• {s}" for s in rat.get("strengths", [])])
        gaps = "\n".join([f"• {g}" for g in rat.get("gaps_and_risks", [])])
        flags = "\n".join([f"• {f}" for f in rat.get("concerns_flags", [])]) if rat.get("concerns_flags") else "None"
        
        # Experience timeline summary
        exp_list = snap.get("experience", [])
        if exp_list:
            exp_summary = " -> ".join([job.get("title", "Unknown") for job in reversed(exp_list)])
        else:
            exp_summary = "No work history listed"
            
        # Inferred skills
        inferred_skills = ", ".join(snap.get("inferred_skills", [])) if snap.get("inferred_skills") else "None"
        
        data.append({
            "Rank": rank + 1,
            "Candidate Name": cand.get("name", "Unknown"),
            "Overall Match Score (%)": cand.get("overall_score", 0),
            "Semantic Alignment Score (%)": sub.get("semantic_score", 0),
            "Hard Skills Overlap (%)": sub.get("skills_score", 0),
            "Experience Fit (%)": sub.get("experience_score", 0),
            "Trajectory & Stability (%)": sub.get("trajectory_score", 0),
            "Executive Summary": rat.get("overall_rank_summary", "N/A"),
            "Grounded Strengths": strengths,
            "Gaps & Risks": gaps,
            "Recruiter Alerts / Flags": flags,
            "Experience Timeline": exp_summary,
            "Total Tenure (Years)": snap.get("overall_tenure_years", 0.0),
            "Avg Job Stay (Years)": snap.get("avg_tenure_years", 0.0),
            "Career Progression": snap.get("growth_trajectory", "stable"),
            "Inferred Skills": inferred_skills
        })
        
    df = pd.DataFrame(data)
    
    # Write to Excel Bytes buffer
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Candidate Rankings")
        
        workbook = writer.book
        worksheet = writer.sheets["Candidate Rankings"]
        
        # Header formatting
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Sleek charcoal gray header
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Cell borders and alignment
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        cell_font = Font(name="Segoe UI", size=10)
        
        for row in range(2, len(data) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = thin_border
                
                # Align scores & rank to center, others to left
                if col in [1, 3, 4, 5, 6, 7, 13, 14]:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    
        # Set dynamic column widths with a max constraint
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                # Split by newline to check line length for wrapped text
                lines = val.split('\n')
                longest_line = max(len(l) for l in lines) if lines else 0
                max_len = max(max_len, longest_line)
            # Add padding and limit to max 50 width, min 10
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
    buffer.seek(0)
    return buffer

# Page Configuration for premium branding

st.set_page_config(
    page_title="DeepMatch - AI Candidate Shortlist Portal",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom premium CSS injection
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&family=Plus+Jakarta+Sans:wght@300;400;600;700&display=swap');
        
        /* Globals */
        html, body, [data-testid="stAppViewContainer"] {
            font-family: 'Plus Jakarta Sans', sans-serif;
            background-color: #090d16;
            color: #f3f4f6;
        }
        
        h1, h2, h3, h4, h5, h6 {
            font-family: 'Outfit', sans-serif !important;
            font-weight: 700 !important;
        }
        
        /* Sleek card styling */
        .candidate-box {
            background-color: #111827;
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 16px;
            padding: 1.5rem;
            margin-bottom: 1.5rem;
            transition: border-color 0.2s ease, transform 0.2s ease;
        }
        .candidate-box:hover {
            border-color: rgba(99, 102, 241, 0.4);
            transform: translateY(-2px);
        }
        
        /* Lists */
        .strength-item {
            color: #10b981;
            font-size: 0.9rem;
            margin-bottom: 0.4rem;
        }
        .gap-item {
            color: #f59e0b;
            font-size: 0.9rem;
            margin-bottom: 0.4rem;
        }
        .flag-box {
            background-color: rgba(239, 68, 68, 0.05);
            border: 1px solid rgba(239, 68, 68, 0.2);
            border-radius: 8px;
            padding: 0.75rem;
            margin-top: 1rem;
            color: #fca5a5;
        }
        
        /* Subtitle */
        .text-muted {
            color: #9ca3af;
        }
        
        /* Sidebar styling override */
        [data-testid="stSidebar"] {
            background-color: #0d121f !important;
            border-right: 1px solid rgba(255, 255, 255, 0.05);
        }
    </style>
""", unsafe_allow_html=True)

# App Title & Branding Header
st.markdown("### 🎯 DEEPMATCH COGNITIVE RECRUITING")
st.title("AI-Powered Candidate Shortlist Portal")
st.markdown("<p class='text-muted'>Explore semantic job matching, dynamic weight tuning, and evidence-grounded explainable rationales.</p>", unsafe_allow_html=True)

# --- Upload Center ---
st.markdown("### 📤 Upload Center")
col_up1, col_up2 = st.columns([1.2, 2])
with col_up1:
    uploaded_jd = st.file_uploader("Upload Job Description (PDF, DOCX, TXT)", type=["pdf", "docx", "txt"])
with col_up2:
    uploaded_resumes = st.file_uploader("Upload Candidate Resumes (Up to 10 files)", type=["pdf", "docx", "txt"], accept_multiple_files=True)

# Process Button & Trigger logic
if uploaded_jd and uploaded_resumes:
    # Check if this combination of files is already processed
    uploaded_files_hash = f"{uploaded_jd.name}_{uploaded_jd.size}_" + "_".join([f"{f.name}_{f.size}" for f in uploaded_resumes])
    
    # Show process button if it's new or they want to re-run
    if st.button("🚀 Process and Rank Uploaded Candidates"):
        with st.spinner("Extracting text and analyzing profiles via Groq API..."):
            try:
                # 1. Extract and parse JD
                jd_bytes = uploaded_jd.read()
                jd_text = extract_text_from_file(jd_bytes, uploaded_jd.name)
                jd_intent = parse_job_description(jd_text)
                
                # 2. Extract and parse resumes
                shortlisted = []
                for res_file in uploaded_resumes[:10]: # Limit to 10 resumes
                    res_bytes = res_file.read()
                    res_text = extract_text_from_file(res_bytes, res_file.name)
                    
                    # LLM structure extraction
                    raw_dict = parse_raw_resume_text(res_text)
                    
                    # Synthesize career metrics and details
                    profile = synthesize_candidate_profile(raw_dict)
                    
                    # Compute scores
                    score, sub_scores, flags = score_candidate(jd_intent, profile)
                    
                    # Generate explainable rationale
                    rationale = generate_candidate_rationale(jd_intent, profile, sub_scores, flags)
                    
                    shortlisted.append({
                        "candidate_id": profile.candidate_id,
                        "name": profile.name,
                        "overall_score": score,
                        "sub_scores": sub_scores.model_dump(),
                        "rationale": rationale.model_dump(),
                        "candidate_profile_snapshot": profile.model_dump()
                    })
                
                # Sort descending
                shortlisted.sort(key=lambda x: x["overall_score"], reverse=True)
                
                # Cache results in Session State
                st.session_state["jd_intent"] = jd_intent.model_dump()
                st.session_state["shortlist"] = shortlisted
                st.session_state["processed_files_hash"] = uploaded_files_hash
                st.success(f"Successfully processed {len(shortlisted)} resumes against the uploaded Job Description!")
                
            except Exception as e:
                st.error(f"Error during file processing: {e}")

# Session state initialization with pre-computed scenario if no uploads are active
if "shortlist" not in st.session_state:
    JSON_PATH = "shortlist.json"
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            st.session_state["jd_intent"] = data["jd_intent"]
            st.session_state["shortlist"] = data["shortlist"]
            st.session_state["processed_files_hash"] = "mock_default"
    else:
        st.info("💡 Upload a Job Description and at least one Resume above to generate a candidate shortlist.")
        st.stop()

# Ensure we have active data from session state
jd = st.session_state["jd_intent"]
original_shortlist = st.session_state["shortlist"]

if not original_shortlist:
    st.warning("No candidates found in the current shortlist. Please upload valid resume files.")
    st.stop()

# Sidebar settings for sliders
st.sidebar.markdown("## ⚙️ Scoring Weight Tuning")
st.sidebar.markdown("Adjust candidate evaluation parameters below. Rankings will update instantly.")

w_semantic = st.sidebar.slider("Semantic Relevance Weight", 0.0, 1.0, 0.30, 0.05)
w_skills = st.sidebar.slider("Hard Skills Overlap", 0.0, 1.0, 0.30, 0.05)
w_exp = st.sidebar.slider("Seniority & Experience Fit", 0.0, 1.0, 0.20, 0.05)
w_traj = st.sidebar.slider("Trajectory & Stability", 0.0, 1.0, 0.20, 0.05)

# Normalize weights so they sum to 1.0
total_w = w_semantic + w_skills + w_exp + w_traj
if total_w > 0:
    n_semantic = w_semantic / total_w
    n_skills = w_skills / total_w
    n_exp = w_exp / total_w
    n_traj = w_traj / total_w
else:
    n_semantic, n_skills, n_exp, n_traj = 0.25, 0.25, 0.25, 0.25

st.sidebar.markdown("---")
st.sidebar.markdown(f"**Normalized Weights Applied:**")
st.sidebar.write(f"- Semantic Match: `{n_semantic:.1%}`")
st.sidebar.write(f"- Skills Overlap: `{n_skills:.1%}`")
st.sidebar.write(f"- Experience Fit: `{n_exp:.1%}`")
st.sidebar.write(f"- Trajectory/Stability: `{n_traj:.1%}`")

# Recalculate candidate scores programmatically on slider inputs (no LLM calls)
ranked_list = []
for c in original_shortlist:
    sub = c["sub_scores"]
    new_overall = (
        (sub["semantic_score"] * n_semantic) +
        (sub["skills_score"] * n_skills) +
        (sub["experience_score"] * n_exp) +
        (sub["trajectory_score"] * n_traj)
    )
    # Clone and write updated overall
    updated_c = dict(c)
    updated_c["overall_score"] = round(new_overall, 2)
    ranked_list.append(updated_c)

# Re-sort list based on updated scores
ranked_list.sort(key=lambda x: x["overall_score"], reverse=True)

# Layout: Main Grid (Left Sidebar with JD intent, Right pane with Candidates)
col_jd, col_cands = st.columns([1, 2.5])

with col_jd:
    st.markdown("### 📋 Target Job Profile")
    st.markdown(f"**Role Title:** `{jd['role_title']}`")
    st.markdown(f"**Target Seniority:** `{jd['seniority_level']}`")
    st.markdown(f"**Target Domain:** `{jd['domain']}`")
    st.markdown(f"**Experience Requirement:** `{jd['min_years_experience']} years`" if jd.get('min_years_experience') else "**Experience Requirement:** `Not specified`")
    
    st.markdown("#### Must-Have Skills")
    st.write(", ".join([f"`{s}`" for s in jd["must_have_skills"]]) if jd.get("must_have_skills") else "None")
    
    st.markdown("#### Nice-To-Have Skills")
    st.write(", ".join([f"`{s}`" for s in jd["nice_to_have_skills"]]) if jd.get("nice_to_have_skills") else "None")
    
    st.markdown("#### Implied Soft Signals")
    st.write(", ".join([f"`{s}`" for s in jd["implied_soft_skills"]]) if jd.get("implied_soft_skills") else "None")
    
    st.markdown("---")
    st.markdown("#### Mission Summary")
    st.write(jd["summary"])

with col_cands:
    col_cands_title, col_download = st.columns([1.8, 1])
    with col_cands_title:
        st.markdown("### 👥 Evaluated Candidates Shortlist")
    with col_download:
        excel_data = generate_excel_report(ranked_list)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name=f"DeepMatch_Report_{jd.get('role_title', 'Shortlist').replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    for rank, cand in enumerate(ranked_list):
        sub = cand["sub_scores"]
        rat = cand["rationale"]
        snap = cand["candidate_profile_snapshot"]
        
        # Color coding for candidate badge
        score = cand["overall_score"]
        if score >= 75:
            score_color = "#10b981"
        elif score >= 55:
            score_color = "#f59e0b"
        else:
            score_color = "#ef4444"
            
        exp_header = " &rarr; ".join([job["title"] for job in reversed(snap["experience"])]) if snap.get("experience") else "No work history listed"
        
        # Candidate layout box
        with st.container():
            st.markdown(f"""
                <div class="candidate-box">
                    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 1rem;">
                        <div>
                            <span style="background-color: rgba(99, 102, 241, 0.15); color: #818cf8; font-weight: bold; padding: 0.2rem 0.6rem; border-radius: 4px; font-size: 0.8rem; margin-right: 0.5rem;">RANK #{rank+1}</span>
                            <span style="font-size: 1.4rem; font-weight: 700; font-family: 'Outfit', sans-serif;">{cand['name']}</span>
                            <div style="font-size: 0.85rem; color: #9ca3af; margin-top: 0.2rem;">{exp_header}</div>
                        </div>
                        <div style="font-family: 'Outfit', sans-serif; font-size: 1.6rem; font-weight: 700; border: 2px solid {score_color}; color: {score_color}; background-color: rgba(255,255,255,0.02); width: 65px; height: 65px; border-radius: 50%; display: flex; align-items: center; justify-content: center;">
                            {score:.0f}
                        </div>
                    </div>
                </div>
            """, unsafe_allow_html=True)
            
            # Subscore breakdown progress bars
            col_b1, col_b2 = st.columns(2)
            with col_b1:
                st.markdown(f"**Semantic Alignment**: `{sub['semantic_score']:.0f}%`")
                st.progress(float(sub["semantic_score"]) / 100.0)
                
                st.markdown(f"**Hard Skills Overlap**: `{sub['skills_score']:.0f}%`")
                st.progress(float(sub["skills_score"]) / 100.0)
                
            with col_b2:
                st.markdown(f"**Experience/Seniority Fit**: `{sub['experience_score']:.0f}%`")
                st.progress(float(sub["experience_score"]) / 100.0)
                
                st.markdown(f"**Career Trajectory/Stability**: `{sub['trajectory_score']:.0f}%`")
                st.progress(float(sub["trajectory_score"]) / 100.0)
                
            # Recruiter explanations
            st.markdown(f"***Executive Summary:*** *\"{rat['overall_rank_summary']}\"*")
            
            col_l, col_r = st.columns(2)
            with col_l:
                st.markdown("##### ✅ Grounded Strengths")
                for s in rat["strengths"]:
                    st.markdown(f"<div class='strength-item'>• {s}</div>", unsafe_allow_html=True)
                    
            with col_r:
                st.markdown("##### ⚠️ Gaps & Risks")
                for g in rat["gaps_and_risks"]:
                    st.markdown(f"<div class='gap-item'>• {g}</div>", unsafe_allow_html=True)
                    
            # Recruiter Concerns & Flags
            if rat["concerns_flags"]:
                st.markdown(f"""
                    <div class="flag-box">
                        <strong>🚩 Recruiter Alerts:</strong>
                        <ul style="margin-top: 0.4rem; padding-left: 1.2rem; margin-bottom: 0;">
                            {"".join([f"<li>{f}</li>" for f in rat["concerns_flags"]])}
                        </ul>
                    </div>
                """, unsafe_allow_html=True)
                
            # Expandable complete resume details
            with st.expander("📄 View Career Timeline & Platform Details"):
                col_exp_1, col_exp_2 = st.columns([2, 1])
                with col_exp_1:
                    st.markdown("#### Work History Details")
                    if snap.get("experience"):
                        for job in snap["experience"]:
                            st.markdown(f"**{job['title']}** at `{job['company']}` ({job['tenure_years']} years)")
                            st.markdown(f"<span style='color: #9ca3af; font-size: 0.9rem;'>{job['description']}</span>", unsafe_allow_html=True)
                            st.write(f"- *Ownership:* {job.get('scope_of_ownership', 'N/A')}")
                            st.write(f"- *Achievement Language:* {job.get('achievement_language', 'N/A')}")
                            st.write("---")
                    else:
                        st.write("No specific work history experiences detailed.")
                with col_exp_2:
                    st.markdown("#### Career Trajectory")
                    st.write(f"- Total Tenure: `{snap.get('overall_tenure_years', 0.0)} years`")
                    st.write(f"- Avg Job Stay: `{snap.get('avg_tenure_years', 0.0)} years`")
                    st.write(f"- Career Progression: `{snap.get('growth_trajectory', 'stable')}`")
                    
                    st.markdown("#### Inferred Skillset")
                    st.write(", ".join([f"`{s}`" for s in snap["inferred_skills"]]) if snap.get("inferred_skills") else "None")
                    
                    if snap.get("platform_activity"):
                        st.markdown("#### Platform Activity")
                        pa = snap["platform_activity"]
                        st.write(f"- Github Commits (Yr): `{pa.get('github_commits_past_year') or 'N/A'}`")
                        st.write(f"- Code Relevance Score: `{pa.get('github_repo_relevance') or 'N/A'}`")
                        st.write(f"- LinkedIn Recency: `{pa.get('linkedin_activity_recency') or 'N/A'}`")
                        st.write(f"- StackOverflow Rep: `{pa.get('stackoverflow_reputation') or 'N/A'}`")
                        
            st.markdown("<br><hr style='border: 1px solid rgba(255,255,255,0.05);'><br>", unsafe_allow_html=True)
